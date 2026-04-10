# ============================================================
# mapping_coverage 필터링 코드 v2 (Google Colab용)
#
# 조건:
#   1. B열(number) 중복 시 → C열이 A/T/G/C 인 행 우선 선택 (1개)
#   2. H열(total) >= 100 인 행만 유지
#   3. I, J, K, L 중 0.05 이상 AND 0.5 미만 인 값이
#      하나라도 있는 행만 유지
#   4. Coverage = H>=100 통과 행 수 / 29903 × 100 (%)
#      (전체 reference position 기준)
#   → 모든 탭 별도 처리 후 하나의 Excel 파일로 저장
# ============================================================

# ── 필요 라이브러리 설치 (Colab 환경) ──────────────────────
# !pip install openpyxl

from google.colab import files
import openpyxl
from io import BytesIO
from collections import defaultdict

REFERENCE_LENGTH = 29903   # SARS-CoV-2 reference genome 전체 position 수

# ── 1. 파일 업로드 (filechooser) ───────────────────────────
print("▶ Excel 파일을 선택하세요 (.xlsx)")
uploaded = files.upload()
filename  = list(uploaded.keys())[0]
file_bytes = BytesIO(uploaded[filename])

# ── 2. 데이터 로드 (수식 계산값 읽기) ──────────────────────
print(f"\n✔ '{filename}' 로드 중...")
wb = openpyxl.load_workbook(file_bytes, data_only=True)
sheet_names = wb.sheetnames
print(f"  발견된 탭 ({len(sheet_names)}개): {sheet_names}\n")

# ── 3. 탭별 필터링 ─────────────────────────────────────────
# 컬럼 인덱스 (1-based, openpyxl 기준)
# A=1(genome ID), B=2(number/position), C=3(ref base),
# D~G=4~7(A,C,G,T counts), H=8(total), I~L=9~12(A,C,G,T ratio)
COL_B = 2   # position number
COL_C = 3   # reference base (A/T/G/C/-)
COL_H = 8   # total read depth
COL_I = 9   # A ratio
COL_J = 10  # C ratio
COL_K = 11  # G ratio
COL_L = 12  # T ratio

VALID_BASES = {'A', 'T', 'G', 'C'}

def safe_float(v):
    """None 또는 변환 불가 값은 -1 반환 (조건 미충족 처리)"""
    try:
        return float(v) if v is not None else -1.0
    except (TypeError, ValueError):
        return -1.0

def has_minor_variant(i_val, j_val, k_val, l_val):
    """I~L 중 0.05 이상 AND 0.5 미만인 값이 하나라도 있으면 True"""
    return any(0.05 <= safe_float(v) < 0.5 for v in [i_val, j_val, k_val, l_val])

summary_rows = []
output_wb    = openpyxl.Workbook()
output_wb.remove(output_wb.active)

for sheet_name in sheet_names:
    ws      = wb[sheet_name]
    max_row = ws.max_row
    max_col = ws.max_column
    headers = [ws.cell(1, c).value for c in range(1, max_col + 1)]

    # ── Step 1: B열 중복 처리 ─────────────────────────────
    # B값별로 행 인덱스 수집 후, A/T/G/C 행 우선 선택
    b_rows = defaultdict(list)   # {b_val: [(row_idx, c_val), ...]}

    for r in range(2, max_row + 1):
        b_val = ws.cell(r, COL_B).value
        c_val = ws.cell(r, COL_C).value
        if b_val is not None:
            b_rows[b_val].append((r, c_val))

    # 각 position에서 대표 행 1개 선택
    # 우선순위: A/T/G/C 인 행 → 없으면 첫 번째 행
    selected_rows = []
    for b_val, row_list in b_rows.items():
        chosen = None
        for (r, c_val) in row_list:
            if c_val in VALID_BASES:
                chosen = r
                break
        if chosen is None:
            chosen = row_list[0][0]   # fallback: 첫 번째 행
        selected_rows.append(chosen)

    selected_rows.sort()  # 행 순서 유지

    # ── Step 2: H>=100 필터 + coverage 계산 ──────────────
    total_positions   = len(selected_rows)          # 중복 제거 후 position 수
    positions_over100 = 0
    filtered_rows     = []

    for r in selected_rows:
        h_val = safe_float(ws.cell(r, COL_H).value)
        if h_val < 0:
            continue

        if h_val >= 100:
            positions_over100 += 1

            # ── Step 3: I~L 중 0.05이상 0.5미만 조건 ────
            if has_minor_variant(
                ws.cell(r, COL_I).value,
                ws.cell(r, COL_J).value,
                ws.cell(r, COL_K).value,
                ws.cell(r, COL_L).value
            ):
                row_data = [ws.cell(r, c).value for c in range(1, max_col + 1)]
                filtered_rows.append(row_data)

    # Coverage = H>=100 인 position 수 / 29903 × 100
    coverage_pct = positions_over100 / REFERENCE_LENGTH * 100

    print(f"[{sheet_name}]")
    print(f"  중복 제거 후 position 수          : {total_positions:,}")
    print(f"  H>=100 position 수 (coverage 분자): {positions_over100:,}")
    print(f"  Coverage (vs 29903)               : {coverage_pct:.2f}%")
    print(f"  최종 필터 통과 행 수               : {len(filtered_rows):,}")
    print()

    summary_rows.append({
        "Sheet"                    : sheet_name,
        "중복제거_position수"       : total_positions,
        "H>=100_position수"        : positions_over100,
        "Coverage_vs29903(%)"      : round(coverage_pct, 2),
        "최종_필터_통과_행수"       : len(filtered_rows),
    })

    # ── 필터 결과를 출력 워크북에 저장 ───────────────────
    new_ws = output_wb.create_sheet(title=sheet_name[:31])
    new_ws.append(headers)
    for row_data in filtered_rows:
        new_ws.append(row_data)

# ── 4. Summary 시트 추가 (맨 앞) ───────────────────────────
summary_ws = output_wb.create_sheet(title="Summary", index=0)
if summary_rows:
    summary_ws.append(list(summary_rows[0].keys()))
    for row in summary_rows:
        summary_ws.append(list(row.values()))

# 열 너비 자동 조정
for col in summary_ws.columns:
    width = max((len(str(cell.value)) if cell.value else 0) for cell in col)
    summary_ws.column_dimensions[col[0].column_letter].width = width + 4

# ── 5. 저장 및 다운로드 ────────────────────────────────────
output_filename = filename.replace(".xlsx", "_filtered_v2.xlsx")
out_buf = BytesIO()
output_wb.save(out_buf)
out_buf.seek(0)

with open(output_filename, "wb") as f:
    f.write(out_buf.read())

print("=" * 60)
print(f"✅ 완료! 결과 파일: {output_filename}")
print("=" * 60)
print("\n[필터 조건 요약]")
print("  ① B열 중복 → C열이 A/T/G/C 인 행 우선 선택 (1 position = 1 행)")
print("  ② H열(total read depth) >= 100")
print(f"  ③ I/J/K/L 중 하나라도 0.05 이상 AND 0.5 미만")
print(f"  Coverage 분모: {REFERENCE_LENGTH} (SARS-CoV-2 genome 전체 position)")

files.download(output_filename)