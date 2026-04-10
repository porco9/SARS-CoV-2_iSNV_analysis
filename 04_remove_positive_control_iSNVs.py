# ============================================================
# iSNV 필터링: 양성대조군(pos) 샘플 기반 제거 후 pos 컬럼 삭제 (수정판)
# ============================================================
# 수정 내용:
#   - A1셀 "Position"이 pos로 오인되는 버그 수정
#     → col 1은 항상 genome position 열로 고정, 샘플 탐색은 col 2부터
#   - merged cell을 고려한 안전한 컬럼 삭제 처리
# ============================================================

from google.colab import files
import openpyxl
import io

# ── 1. 파일 업로드 ──────────────────────────────────────────
print("▶ 파일을 업로드하세요 (.xlsx)")
uploaded = files.upload()
filename = list(uploaded.keys())[0]
wb = openpyxl.load_workbook(io.BytesIO(uploaded[filename]))
ws = wb.active
print(f"  시트: {ws.title}  |  크기: {ws.dimensions}")

# ── 2. 헤더 파싱 ────────────────────────────────────────────
# Row 1: col1 = "Position" (genome pos 열), col2부터 샘플명 (2열씩)
# Row 2: Major / Minor 반복
# Row 3+: 데이터

row1_values = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

# ★ col 1은 항상 genome position 열 → 샘플 탐색은 col 2부터
# pos 샘플: 샘플명(col1 제외)에 'pos'가 포함된 것
pos_name_cols = [
    c for c, val in enumerate(row1_values, start=1)
    if c >= 2 and val and 'pos' in str(val).lower()
]
pos_minor_cols = [c + 1 for c in pos_name_cols]  # Minor 컬럼 = 샘플명 컬럼 + 1

print(f"\n  양성대조군 샘플 수: {len(pos_name_cols)}")
for c in pos_name_cols:
    print(f"    col {c}: {ws.cell(row=1, column=c).value}")

# ── 3. pos에서 Minor allele가 관찰된 행 식별 ────────────────
pos_isnv_rows = set()
for r in range(3, ws.max_row + 1):
    for mc in pos_minor_cols:
        if ws.cell(row=r, column=mc).value is not None:
            pos_isnv_rows.add(r)
            break

print(f"\n  pos iSNV가 관찰된 genome position 수: {len(pos_isnv_rows)}")

# ── 4. 나머지 샘플에서 해당 행 데이터 제거 ─────────────────
pos_all_cols = set()
for c in pos_name_cols:
    pos_all_cols.add(c)       # Major
    pos_all_cols.add(c + 1)   # Minor

# col 1(genome pos)과 pos 컬럼을 제외한 나머지 데이터 컬럼
non_pos_data_cols = [
    c for c in range(2, ws.max_column + 1)
    if c not in pos_all_cols
]

cleared = 0
for r in pos_isnv_rows:
    for c in non_pos_data_cols:
        if ws.cell(row=r, column=c).value is not None:
            ws.cell(row=r, column=c).value = None
            cleared += 1

print(f"  나머지 샘플에서 제거된 셀 수: {cleared}")

# ── 5. merged cell 해제 후 pos 컬럼 삭제 ────────────────────
# merged cell이 있으면 delete_cols 시 오류 발생 가능 → 먼저 unmerge
for mc in list(ws.merged_cells.ranges):
    ws.unmerge_cells(str(mc))

# pos 컬럼 삭제 (뒤에서부터 삭제해야 인덱스 밀림 방지)
cols_to_delete = sorted(pos_all_cols, reverse=True)
for c in cols_to_delete:
    ws.delete_cols(c)

print(f"\n  삭제된 pos 컬럼 수: {len(cols_to_delete)}")
print(f"  삭제 후 컬럼 수: {ws.max_column}")

# ── 6. 헤더 확인 출력 ───────────────────────────────────────
print("\n  삭제 후 Row1 샘플명 (처음 10열):")
for c in range(1, min(11, ws.max_column + 1)):
    print(f"    col{c}: {ws.cell(1, c).value}")

print("  삭제 후 Row2 헤더 (처음 10열):")
for c in range(1, min(11, ws.max_column + 1)):
    print(f"    col{c}: {ws.cell(2, c).value}")

# ── 7. 데이터가 완전히 비어있는 행 삭제 ─────────────────────
rows_to_delete = []
for r in range(3, ws.max_row + 1):
    all_empty = all(
        ws.cell(row=r, column=c).value is None
        for c in range(2, ws.max_column + 1)
    )
    if all_empty:
        rows_to_delete.append(r)

for r in sorted(rows_to_delete, reverse=True):
    ws.delete_rows(r)

print(f"\n  삭제된 빈 행 수: {len(rows_to_delete)}")
print(f"  최종 데이터 행 수: {ws.max_row - 2}")  # 헤더 2행 제외

# ── 8. 저장 및 다운로드 ─────────────────────────────────────
out_filename = filename.replace('.xlsx', '_filtered.xlsx')
output = io.BytesIO()
wb.save(output)
output.seek(0)

with open(out_filename, 'wb') as f:
    f.write(output.read())

files.download(out_filename)
print(f"\n✅ 완료! 저장 파일: {out_filename}")