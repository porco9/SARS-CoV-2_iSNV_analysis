# ================================================================
# iSNV 단백질별 분류 + 카운트 정리 (2개 파일 생성)
# ================================================================
# 수정: col1 'Position' 이 샘플명으로 오인되는 버그 수정
#       → 샘플 탐색을 col 2부터 시작하도록 변경
# ================================================================

from google.colab import files
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io, re

# ── 파일 업로드: iSNV summary ──────────────────────────────────
print("▶ [1/2] iSNV summary 파일을 업로드하세요 (pos 제거된 파일)")
up1 = files.upload()
isnv_file = list(up1.keys())[0]
print(f"  ✔ iSNV 파일: {isnv_file}")

# ── 파일 업로드: 단백질 position ───────────────────────────────
print("\n▶ [2/2] 단백질 position 파일을 업로드하세요 (SARS_CoV2_protein_position)")
up2 = files.upload()
protein_file = list(up2.keys())[0]
print(f"  ✔ Protein 파일: {protein_file}")

# ── 1. 단백질 position 로드 ────────────────────────────────────
wb_prot = openpyxl.load_workbook(io.BytesIO(up2[protein_file]))
ws_prot = wb_prot.active

TARGET_PROTEINS = ['Spike', 'ORF1a', 'ORF1b', 'ORF3a', 'ORF6', 'ORF8', 'ORF10']

protein_ranges = {}
for row in ws_prot.iter_rows(values_only=True, min_row=2):
    name = str(row[0]).strip() if row[0] else ''
    if name in TARGET_PROTEINS:
        protein_ranges[name] = (int(row[1]), int(row[2]))

print("\n  단백질 범위:")
for p in TARGET_PROTEINS:
    print(f"    {p}: {protein_ranges[p]}")

def get_protein(pos):
    for pname in TARGET_PROTEINS:
        start, stop = protein_ranges[pname]
        if start <= pos <= stop:
            return pname
    return None

# ── 2. iSNV summary 로드 ──────────────────────────────────────
wb_isnv = openpyxl.load_workbook(io.BytesIO(up1[isnv_file]))
ws_isnv = wb_isnv.active

row1_vals = [ws_isnv.cell(1, c).value for c in range(1, ws_isnv.max_column + 1)]

# ★ 버그 수정: c_idx >= 2 조건 추가 → col1 'Position' 제외
sample_cols = []
for c_idx, val in enumerate(row1_vals, start=1):
    if c_idx >= 2 and val is not None:
        sample_cols.append({
            'name': val,
            'major_col': c_idx,
            'minor_col': c_idx + 1
        })

print(f"\n  샘플 수: {len(sample_cols)}")
for sc in sample_cols:
    print(f"    col{sc['major_col']}: {sc['name']}")

# 데이터 수집: {genome_pos: {sample_name: (major, minor)}}
data_by_pos = {}
for r in range(3, ws_isnv.max_row + 1):
    pos = ws_isnv.cell(r, 1).value
    if pos is None:
        continue
    row_data = {}
    for sc in sample_cols:
        maj = ws_isnv.cell(r, sc['major_col']).value
        mino = ws_isnv.cell(r, sc['minor_col']).value
        if maj is not None or mino is not None:
            row_data[sc['name']] = (maj, mino)
    if row_data:
        data_by_pos[int(pos)] = row_data

print(f"  데이터 있는 genome positions 수: {len(data_by_pos)}")

# ── 3. 단백질별로 position 분류 ────────────────────────────────
protein_positions = {p: [] for p in TARGET_PROTEINS}
unclassified = []

for pos in sorted(data_by_pos.keys()):
    pname = get_protein(pos)
    if pname:
        protein_positions[pname].append(pos)
    else:
        unclassified.append(pos)

for p in TARGET_PROTEINS:
    print(f"  {p}: {len(protein_positions[p])} positions")
if unclassified:
    print(f"  미분류: {unclassified}")

# ── 스타일 헬퍼 ───────────────────────────────────────────────
def header_fill(color):
    return PatternFill("solid", fgColor=color)

PROTEIN_COLORS = {
    'Spike':  'D9E1F2',
    'ORF1a':  'E2EFDA',
    'ORF1b':  'FFF2CC',
    'ORF3a':  'FCE4D6',
    'ORF6':   'EDEDED',
    'ORF8':   'F4CCFF',
    'ORF10':  'CCF4FF',
}
THIN = Side(style='thin')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def set_header_cell(ws, row, col, value, bold=True, fill_color=None, center=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold)
    if fill_color:
        cell.fill = header_fill(fill_color)
    if center:
        cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = BORDER
    return cell

# ── 4. 파일 1: 단백질별 시트로 분리된 iSNV summary ─────────────
print("\n▶ 파일 1 생성 중: 단백질별 iSNV summary...")

wb_out1 = openpyxl.Workbook()
wb_out1.remove(wb_out1.active)

PROTEIN_COLORS['Other'] = 'F2F2F2'
if unclassified:
    protein_positions['Other'] = sorted(unclassified)
    print(f"  Other (미분류): {len(unclassified)} positions → 'Other' 시트로 저장")
ALL_SHEETS = TARGET_PROTEINS + (['Other'] if unclassified else [])

for pname in ALL_SHEETS:
    positions = sorted(protein_positions.get(pname, []))

    ws_new = wb_out1.create_sheet(title=pname)
    color = PROTEIN_COLORS.get(pname, 'FFFFFF')

    ws_new.cell(1, 1, 'Sample').font = Font(bold=True)
    ws_new.cell(1, 1).fill = header_fill('BFBFBF')
    ws_new.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')

    ws_new.cell(2, 1, 'Reference').font = Font(bold=True)
    ws_new.cell(2, 1).fill = header_fill('BFBFBF')
    ws_new.cell(2, 1).alignment = Alignment(horizontal='center', vertical='center')

    for s_idx, sc in enumerate(sample_cols):
        maj_col = 2 + s_idx * 2
        min_col = maj_col + 1

        ws_new.cell(1, maj_col, sc['name'])
        ws_new.merge_cells(
            start_row=1, start_column=maj_col,
            end_row=1, end_column=min_col
        )
        cell = ws_new.cell(1, maj_col)
        cell.font = Font(bold=True)
        cell.fill = header_fill(color)
        cell.alignment = Alignment(horizontal='center', vertical='center')

        set_header_cell(ws_new, 2, maj_col, 'Major', fill_color=color)
        set_header_cell(ws_new, 2, min_col, 'Minor', fill_color=color)

    for r_offset, pos in enumerate(positions):
        data_row = 3 + r_offset
        ws_new.cell(data_row, 1, pos)
        ws_new.cell(data_row, 1).alignment = Alignment(horizontal='center')

        row_data = data_by_pos.get(pos, {})
        for s_idx, sc in enumerate(sample_cols):
            maj_col = 2 + s_idx * 2
            min_col = maj_col + 1
            if sc['name'] in row_data:
                maj, mino = row_data[sc['name']]
                if maj is not None:
                    ws_new.cell(data_row, maj_col, maj)
                if mino is not None:
                    ws_new.cell(data_row, min_col, mino)

    ws_new.column_dimensions['A'].width = 12
    for s_idx in range(len(sample_cols)):
        ws_new.column_dimensions[get_column_letter(2 + s_idx * 2)].width = 14
        ws_new.column_dimensions[get_column_letter(3 + s_idx * 2)].width = 10

    print(f"  [{pname}] {len(positions)} positions × {len(sample_cols)} samples")

# ── 5. 파일 2: 샘플별 단백질별 iSNV count ─────────────────────
print("\n▶ 파일 2 생성 중: 샘플별 iSNV count...")

def count_isnvs_per_sample():
    counts = {}
    for sc in sample_cols:
        counts[sc['name']] = {p: 0 for p in TARGET_PROTEINS}
        counts[sc['name']]['total'] = 0

    for pos, row_data in data_by_pos.items():
        pname = get_protein(pos)
        if pname is None:
            continue
        for sname, (maj, mino) in row_data.items():
            if mino is not None:
                counts[sname][pname] += 1
                counts[sname]['total'] += 1

    return counts

isnv_counts = count_isnvs_per_sample()

wb_out2 = openpyxl.Workbook()
ws_count = wb_out2.active
ws_count.title = 'iSNV_count'

HEADER_COLOR = 'D6DCE4'
headers = ['Sample', 'Group', 'total_iSNVs'] + [f'{p}_iSNVs' for p in TARGET_PROTEINS]
for c_idx, h in enumerate(headers, start=1):
    set_header_cell(ws_count, 1, c_idx, h, fill_color=HEADER_COLOR)

def extract_group(sample_name):
    m = re.search(r'_(\d+\.?\d*)_', sample_name)
    if m:
        return m.group(1)
    m = re.search(r'_([\d.]+)\s', sample_name)
    if m:
        return m.group(1)
    return 'unknown'

def group_sort_key(sname):
    g = extract_group(sname)
    try:
        return float(g)
    except:
        return 999

sorted_samples = sorted(sample_cols, key=lambda sc: group_sort_key(sc['name']))

row_colors = {'0.01': 'EAF4FF', '0.1': 'FFF4EA', '1': 'F4FFEA'}

for r_offset, sc in enumerate(sorted_samples):
    data_row = 2 + r_offset
    sname = sc['name']
    group = extract_group(sname)
    cnt = isnv_counts[sname]

    fill_color = row_colors.get(group, 'FFFFFF')

    row_vals = [sname, group, cnt['total']] + [cnt[p] for p in TARGET_PROTEINS]
    for c_idx, val in enumerate(row_vals, start=1):
        cell = ws_count.cell(data_row, c_idx, val)
        cell.fill = header_fill(fill_color)
        cell.alignment = Alignment(horizontal='center')
        cell.border = BORDER

ws_count.column_dimensions['A'].width = 25
ws_count.column_dimensions['B'].width = 10
for c_idx in range(3, len(headers) + 1):
    ws_count.column_dimensions[get_column_letter(c_idx)].width = 14

print(f"  샘플 {len(sorted_samples)}개 카운트 완료")

# 합계 행
sum_row = 2 + len(sorted_samples)
ws_count.cell(sum_row, 1, 'Sum').font = Font(bold=True)
ws_count.cell(sum_row, 1).fill = header_fill('BFBFBF')
ws_count.cell(sum_row, 1).alignment = Alignment(horizontal='center')
ws_count.cell(sum_row, 1).border = BORDER

ws_count.cell(sum_row, 2, '').fill = header_fill('BFBFBF')
ws_count.cell(sum_row, 2).border = BORDER

total_sum = sum(isnv_counts[sc['name']]['total'] for sc in sample_cols)
ws_count.cell(sum_row, 3, total_sum).font = Font(bold=True)
ws_count.cell(sum_row, 3).fill = header_fill('BFBFBF')
ws_count.cell(sum_row, 3).alignment = Alignment(horizontal='center')
ws_count.cell(sum_row, 3).border = BORDER

for c_offset, p in enumerate(TARGET_PROTEINS):
    col = 4 + c_offset
    psum = sum(isnv_counts[sc['name']][p] for sc in sample_cols)
    cell = ws_count.cell(sum_row, col, psum)
    cell.font = Font(bold=True)
    cell.fill = header_fill('BFBFBF')
    cell.alignment = Alignment(horizontal='center')
    cell.border = BORDER

# ── 6. 저장 및 다운로드 ───────────────────────────────────────
base_name = re.sub(r'\.xlsx$', '', isnv_file, flags=re.IGNORECASE)

out1_name = base_name + '_단백질별_iSNV_summary.xlsx'
out2_name = base_name + '_단백질별_iSNV_count.xlsx'

for out_name, wb_obj in [(out1_name, wb_out1), (out2_name, wb_out2)]:
    buf = io.BytesIO()
    wb_obj.save(buf)
    buf.seek(0)
    with open(out_name, 'wb') as f:
        f.write(buf.read())
    files.download(out_name)
    print(f"\n✅ 저장: {out_name}")

print("\n🎉 모든 작업 완료!")