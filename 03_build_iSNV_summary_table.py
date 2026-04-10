# ============================================================
# iSNV Summary Table Builder
# Google Colab용 - mapping_coverage_filtered_without_pos.xlsx
# 모든 샘플 탭을 하나의 요약 탭으로 통합
# ============================================================

# ── 0. 라이브러리 설치 ──────────────────────────────────────
# openpyxl은 Colab 기본 포함이지만 명시적으로 확인
# !pip install openpyxl --quiet

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from google.colab import files

# ── 1. 파일 업로드 ────────────────────────────────────────
print("📂 분석할 Excel 파일을 선택하세요 (mapping_coverage_filtered_without_pos.xlsx)")
uploaded = files.upload()
input_filename = list(uploaded.keys())[0]
print(f"✅ 업로드 완료: {input_filename}")

# ── 2. 시트 로드 및 데이터 파싱 ───────────────────────────
xl = pd.ExcelFile(input_filename)
sample_sheets = [s for s in xl.sheet_names if s.lower() != 'summary']
print(f"📋 샘플 시트 수: {len(sample_sheets)}개")
print(f"   {sample_sheets}")

# 염기 및 비율 컬럼 인덱스 (0-based)
# 컬럼 구조: 0=sample_id, 1=number, 2=ref_genome,
#            3~7=A,C,G,T count+total, 8~12=A,C,G,T ratio+total
BASES     = ['A', 'C', 'G', 'T']
RATIO_COLS = [8, 9, 10, 11]   # 비율 컬럼 (I,J,K,L → 0-index 8,9,10,11)

all_positions = set()
sample_data   = {}  # {sample_name: {pos: {'major': str, 'minor': str}}}

for sheet in sample_sheets:
    df = pd.read_excel(xl, sheet_name=sheet, header=None)
    df = df.iloc[1:].reset_index(drop=True)   # 헤더 행 제거
    df.columns = range(df.shape[1])

    info = {}
    for _, row in df.iterrows():
        pos = row[1]
        if pd.isna(pos):
            continue
        pos = int(pos)
        all_positions.add(pos)

        # 각 염기의 비율 수집
        ratios = {}
        for base, col in zip(BASES, RATIO_COLS):
            try:
                v = float(row[col])
                if pd.notna(v):
                    ratios[base] = v
            except (ValueError, TypeError):
                pass

        # Major (>0.5) / Minor (0.05~0.5) 분류
        major = None
        minor = None
        for base, v in ratios.items():
            if v > 0.5:
                major = f"{base} {round(v, 2)}"
            elif 0.05 <= v < 0.5:
                minor = f"{base} {round(v, 2)}"

        if major or minor:
            info[pos] = {'major': major, 'minor': minor}

    sample_data[sheet] = info

sorted_positions = sorted(all_positions)
print(f"\n🔬 총 unique reference positions: {len(sorted_positions)}개")

# ── 3. 스타일 정의 ────────────────────────────────────────
def make_fill(hex_color):
    return PatternFill("solid", start_color=hex_color, end_color=hex_color)

def make_border(color="BFBFBF"):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)

STYLE = {
    "header1_fill" : make_fill("1F4E79"),
    "header1_font" : Font(bold=True, color="FFFFFF", name="Arial", size=10),
    "header2_fill" : make_fill("2E75B6"),
    "header2_font" : Font(bold=True, color="FFFFFF", name="Arial", size=9),
    "pos_fill"     : make_fill("D6E4F0"),
    "pos_font"     : Font(bold=True, name="Arial", size=9),
    "major_fill"   : make_fill("E2EFDA"),
    "major_font"   : Font(name="Arial", size=9, color="375623"),
    "minor_fill"   : make_fill("FFF2CC"),
    "minor_font"   : Font(name="Arial", size=9, color="7F6000"),
    "data_font"    : Font(name="Arial", size=9),
    "center"       : Alignment(horizontal="center", vertical="center", wrap_text=True),
    "border"       : make_border(),
}

# ── 4. 워크북 생성 ────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "iSNV_Summary"

COL_START = 2   # B열부터 샘플 데이터 시작

# ── 4-1. 헤더 행 1,2 작성 ──────────────────────────────
# A열: "Sample" (행 1~2 병합) / "Reference"
ws.merge_cells("A1:A2")
c = ws["A1"]
c.value, c.font, c.fill, c.alignment, c.border = (
    "Sample", STYLE["header1_font"], STYLE["header1_fill"],
    STYLE["center"], STYLE["border"]
)

ws["A2"].border = STYLE["border"]

# B열 이후: 샘플명(1행 병합) + Major/Minor(2행)
for i, sample in enumerate(sample_sheets):
    c1 = COL_START + i * 2
    c2 = c1 + 1
    l1, l2 = get_column_letter(c1), get_column_letter(c2)

    ws.merge_cells(f"{l1}1:{l2}1")
    cell = ws[f"{l1}1"]
    cell.value      = sample
    cell.font       = STYLE["header1_font"]
    cell.fill       = STYLE["header1_fill"]
    cell.alignment  = STYLE["center"]
    cell.border     = STYLE["border"]
    ws[f"{l2}1"].border = STYLE["border"]

    for label, col in [("Major", l1), ("Minor", l2)]:
        cell = ws[f"{col}2"]
        cell.value     = label
        cell.font      = STYLE["header2_font"]
        cell.fill      = STYLE["header2_fill"]
        cell.alignment = STYLE["center"]
        cell.border    = STYLE["border"]

# ── 4-2. Position 및 allele 데이터 작성 ───────────────
for r, pos in enumerate(sorted_positions):
    row_idx = r + 3

    # A열: position
    c = ws.cell(row=row_idx, column=1, value=pos)
    c.font, c.fill, c.alignment, c.border = (
        STYLE["pos_font"], STYLE["pos_fill"], STYLE["center"], STYLE["border"]
    )

    # 각 샘플의 Major / Minor 값 기입
    for i, sample in enumerate(sample_sheets):
        c1 = COL_START + i * 2
        c2 = c1 + 1
        allele = sample_data[sample].get(pos, {})

        for col_idx, key, fill_key, font_key in [
            (c1, "major", "major_fill", "major_font"),
            (c2, "minor", "minor_fill", "minor_font"),
        ]:
            val  = allele.get(key)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = STYLE[font_key]  if val else STYLE["data_font"]
            cell.fill      = STYLE[fill_key]  if val else PatternFill()
            cell.alignment = STYLE["center"]
            cell.border    = STYLE["border"]

# ── 4-3. 열 너비 / 행 높이 / 틀 고정 ────────────────────
ws.column_dimensions['A'].width = 12
for i in range(len(sample_sheets) * 2):
    ws.column_dimensions[get_column_letter(COL_START + i)].width = 10

ws.row_dimensions[1].height = 22
ws.row_dimensions[2].height = 16
ws.freeze_panes = "B3"   # 헤더 + position 열 고정

# ── 5. 저장 및 다운로드 ───────────────────────────────────
output_filename = "iSNV_summary.xlsx"
wb.save(output_filename)
print(f"\n✅ 저장 완료: {output_filename}")
print(f"   - Positions : {len(sorted_positions)}개")
print(f"   - Samples   : {len(sample_sheets)}개")
print(f"   - 총 열 수  : {1 + len(sample_sheets) * 2}열")

files.download(output_filename)
print("📥 다운로드 시작됨")