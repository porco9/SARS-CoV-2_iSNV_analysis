# ================================================================
# SARS-CoV-2 Protein Coverage Analysis (Google Colab)
# H열(total read depth) ≥ 100 기준, 단백질별 Coverage 계산
# 입력: 샘플별 탭이 있는 xlsx 1개 + protein position 파일
#
# [변경사항]
#   Pivot 파일에 Coverage_% 와 Covered_positions(절대수) 모두 표시
#   컬럼 구조: Protein | Start | Stop | Gene_length
#              | Sample1_N | Sample1_% | Sample2_N | Sample2_% | ...
# ================================================================

import openpyxl
import pandas as pd
from google.colab import files
from IPython.display import display
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io, re

READ_DEPTH_THRESHOLD = 100
TOTAL_GENOME_LENGTH  = 29903

# ── 파일 업로드 ──────────────────────────────────────────────
print("▶ [1/2] protein position 파일을 업로드하세요")
print("   (_참고용__SARS_CoV2_protein_position.xlsx)")
up1 = files.upload()
protein_file = list(up1.keys())[0]

wb_prot = openpyxl.load_workbook(io.BytesIO(up1[protein_file]))
ws_prot = wb_prot.active
proteins = []   # [(name, start, stop), ...]
for row in ws_prot.iter_rows(values_only=True, min_row=2):
    if row[0] and row[1] and row[2]:
        proteins.append((str(row[0]).strip(), int(row[1]), int(row[2])))

print(f"  ✔ {len(proteins)}개 단백질 로드 완료\n")

print("▶ [2/2] raw data 파일을 업로드하세요")
print("   (각 샘플이 탭으로 구성된 xlsx, B열=number, H열=total read depth)")
up2 = files.upload()
raw_file = list(up2.keys())[0]
print(f"  ✔ {raw_file} 로드 완료\n")

wb_raw = openpyxl.load_workbook(io.BytesIO(up2[raw_file]))
print(f"  탭(샘플) 수: {len(wb_raw.sheetnames)}개")
print(f"  탭 목록: {wb_raw.sheetnames}\n")

# ================================================================
# Coverage 계산 함수
# ================================================================
def calc_coverage(ws, proteins, threshold=READ_DEPTH_THRESHOLD):
    """
    ws    : openpyxl worksheet (B열=number, H열=total read depth)
    반환  : DataFrame (Protein, Start, Stop, Gene_length,
                       Covered_positions, Coverage_%)
    """
    covered = set()
    for r in range(2, ws.max_row + 1):
        num   = ws.cell(r, 2).value   # B열: genome position
        total = ws.cell(r, 8).value   # H열: total read depth
        if num is None or total is None:
            continue
        try:
            n = int(num)
            t = float(total)
        except (ValueError, TypeError):
            continue
        if t >= threshold:
            covered.add(n)

    results = []

    # 전체 게놈
    wg_cov = len(covered)
    results.append({
        'Protein'           : 'Whole Genome',
        'Start'             : 1,
        'Stop'              : TOTAL_GENOME_LENGTH,
        'Gene_length'       : TOTAL_GENOME_LENGTH,
        'Covered_positions' : wg_cov,
        'Coverage_%'        : round(wg_cov / TOTAL_GENOME_LENGTH * 100, 2)
    })

    # 단백질별
    for pname, start, stop in proteins:
        pset = set(range(start, stop + 1))
        plen = len(pset)
        pcov = len(covered & pset)
        pct  = round(pcov / plen * 100, 2) if plen > 0 else 0.0
        results.append({
            'Protein'           : pname,
            'Start'             : start,
            'Stop'              : stop,
            'Gene_length'       : plen,
            'Covered_positions' : pcov,
            'Coverage_%'        : pct
        })

    return pd.DataFrame(results)

# ================================================================
# 전체 샘플 분석
# ================================================================
print(f"▶ 분석 시작 (H열 total read depth ≥ {READ_DEPTH_THRESHOLD} 기준)\n")
print("=" * 55)

all_results = {}   # {sample_name: DataFrame}

for sheet_name in wb_raw.sheetnames:
    ws = wb_raw[sheet_name]
    print(f"  ⏳ {sheet_name} ...", end=' ')
    try:
        df = calc_coverage(ws, proteins)
        all_results[sheet_name] = df
        wg_row = df[df['Protein'] == 'Whole Genome'].iloc[0]
        print(f"covered {int(wg_row['Covered_positions']):,} / {TOTAL_GENOME_LENGTH:,} "
              f"({wg_row['Coverage_%']:.1f}%)")
    except Exception as e:
        print(f"❌ 오류: {e}")

print()

# ================================================================
# 스타일 유틸
# ================================================================
THIN     = Side(style='thin')
BORDER   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", fgColor="D6DCE4")
WG_FILL  = PatternFill("solid", fgColor="BDD7EE")
N_FILL   = PatternFill("solid", fgColor="EBF5FB")   # 절대수 컬럼 배경 (연파랑)
PCT_FILL = PatternFill("solid", fgColor="FDFEFE")   # % 컬럼 배경 (흰색)

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        cell.font      = Font(bold=True)
        cell.fill      = HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = BORDER

def style_row(ws, row, ncols, fill=None):
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        if fill:
            cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = BORDER

# ================================================================
# 출력 1: 샘플별 개별 시트 (기존과 동일)
# ================================================================
wb_out1 = openpyxl.Workbook()
wb_out1.remove(wb_out1.active)

for sample_name, df in all_results.items():
    ws_new = wb_out1.create_sheet(title=sample_name[:31])
    headers = list(df.columns)
    ncols   = len(headers)

    for c_idx, h in enumerate(headers, start=1):
        ws_new.cell(1, c_idx, h)
    style_header(ws_new, 1, ncols)

    for r_off, (_, row_data) in enumerate(df.iterrows()):
        dr = 2 + r_off
        for c_idx, val in enumerate(row_data.values, start=1):
            ws_new.cell(dr, c_idx, val)
        fill = WG_FILL if row_data['Protein'] == 'Whole Genome' else None
        style_row(ws_new, dr, ncols, fill=fill)

    col_widths = [18, 8, 8, 13, 20, 13]
    for i, w in enumerate(col_widths, start=1):
        ws_new.column_dimensions[get_column_letter(i)].width = w

# ================================================================
# 출력 2: Pivot 테이블
# 컬럼: Protein | Start | Stop | Gene_length
#        | S1_N | S1_% | S2_N | S2_% | ...
# ================================================================
wb_out2 = openpyxl.Workbook()
ws_pivot = wb_out2.active
ws_pivot.title = 'Coverage_Pivot'

sample_names = list(all_results.keys())

# ── 헤더 행 1: 샘플명 병합 ───────────────────────────────────
# 고정 컬럼 4개 + 샘플당 2컬럼(N, %)
FIXED_COLS = 4
fixed_headers = ['Protein', 'Start', 'Stop', 'Gene_length']

# 고정 헤더
for c_idx, h in enumerate(fixed_headers, start=1):
    cell = ws_pivot.cell(1, c_idx, h)
    cell.font      = Font(bold=True)
    cell.fill      = HDR_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border    = BORDER
    # 2행과 병합 (고정 컬럼은 2행까지 병합)
    ws_pivot.merge_cells(
        start_row=1, start_column=c_idx,
        end_row=2,   end_column=c_idx
    )

# 샘플명 헤더 (2컬럼씩 병합)
for s_idx, sname in enumerate(sample_names):
    col_start = FIXED_COLS + 1 + s_idx * 2
    col_end   = col_start + 1
    ws_pivot.merge_cells(
        start_row=1, start_column=col_start,
        end_row=1,   end_column=col_end
    )
    cell = ws_pivot.cell(1, col_start, sname)
    cell.font      = Font(bold=True)
    cell.fill      = HDR_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border    = BORDER

# ── 헤더 행 2: N / % 서브헤더 ────────────────────────────────
for s_idx in range(len(sample_names)):
    col_n   = FIXED_COLS + 1 + s_idx * 2
    col_pct = col_n + 1

    cell_n = ws_pivot.cell(2, col_n, 'Covered_N')
    cell_n.font      = Font(bold=True)
    cell_n.fill      = HDR_FILL
    cell_n.alignment = Alignment(horizontal='center', vertical='center')
    cell_n.border    = BORDER

    cell_pct = ws_pivot.cell(2, col_pct, 'Coverage_%')
    cell_pct.font      = Font(bold=True)
    cell_pct.fill      = HDR_FILL
    cell_pct.alignment = Alignment(horizontal='center', vertical='center')
    cell_pct.border    = BORDER

# ── 데이터 행 ────────────────────────────────────────────────
ncols_total = FIXED_COLS + len(sample_names) * 2

if all_results:
    first_df = next(iter(all_results.values()))
    for r_off, (_, base_row) in enumerate(first_df.iterrows()):
        dr = 3 + r_off   # 헤더 2행이므로 데이터는 3행부터

        # 고정 컬럼
        ws_pivot.cell(dr, 1, base_row['Protein'])
        ws_pivot.cell(dr, 2, base_row['Start'])
        ws_pivot.cell(dr, 3, base_row['Stop'])
        ws_pivot.cell(dr, 4, base_row['Gene_length'])

        # 샘플별 N + %
        for s_idx, sname in enumerate(sample_names):
            col_n   = FIXED_COLS + 1 + s_idx * 2
            col_pct = col_n + 1

            sample_df = all_results[sname]
            match = sample_df[sample_df['Protein'] == base_row['Protein']]

            if len(match) > 0:
                n_val   = int(match['Covered_positions'].values[0])
                pct_val = float(match['Coverage_%'].values[0])
            else:
                n_val   = None
                pct_val = None

            cell_n = ws_pivot.cell(dr, col_n, n_val)
            cell_n.alignment = Alignment(horizontal='center', vertical='center')
            cell_n.border    = BORDER

            cell_pct = ws_pivot.cell(dr, col_pct, pct_val)
            cell_pct.alignment = Alignment(horizontal='center', vertical='center')
            cell_pct.border    = BORDER

            # Whole Genome 행 강조
            if base_row['Protein'] == 'Whole Genome':
                cell_n.fill   = WG_FILL
                cell_pct.fill = WG_FILL
            else:
                cell_n.fill   = N_FILL
                cell_pct.fill = PCT_FILL

        # 고정 컬럼 스타일
        row_fill = WG_FILL if base_row['Protein'] == 'Whole Genome' else None
        for c in range(1, FIXED_COLS + 1):
            cell = ws_pivot.cell(dr, c)
            if row_fill:
                cell.fill = row_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border    = BORDER

# ── 컬럼 너비 ────────────────────────────────────────────────
ws_pivot.column_dimensions['A'].width = 18   # Protein
for c in range(2, 5):
    ws_pivot.column_dimensions[get_column_letter(c)].width = 12

for s_idx, sname in enumerate(sample_names):
    col_n   = FIXED_COLS + 1 + s_idx * 2
    col_pct = col_n + 1
    label_w = max(14, len(sname) + 2)
    ws_pivot.column_dimensions[get_column_letter(col_n)].width   = label_w
    ws_pivot.column_dimensions[get_column_letter(col_pct)].width = label_w

# ── 행 높이 (헤더) ───────────────────────────────────────────
ws_pivot.row_dimensions[1].height = 20
ws_pivot.row_dimensions[2].height = 18

# ================================================================
# 저장 및 다운로드
# ================================================================
print("▶ 파일 저장 및 다운로드 중...")

base = re.sub(r'\.xlsx$', '', raw_file, flags=re.IGNORECASE)
out1_name = base + '_Coverage_by_Sample.xlsx'
out2_name = base + '_Coverage_Pivot.xlsx'

for out_name, wb_obj in [(out1_name, wb_out1), (out2_name, wb_out2)]:
    buf = io.BytesIO()
    wb_obj.save(buf)
    buf.seek(0)
    with open(out_name, 'wb') as f:
        f.write(buf.read())
    files.download(out_name)
    print(f"  ✅ {out_name}")

print("\n🎉 완료!")
print(f"  ├── {out1_name}  (샘플별 시트)")
print(f"  └── {out2_name}  (단백질 × 샘플 Pivot)")
print()
print("  [Pivot 컬럼 구조]")
print("  Protein | Start | Stop | Gene_length")
print("  | Sample1_Covered_N | Sample1_Coverage_% | Sample2_N | Sample2_% | ...")