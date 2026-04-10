# ================================================================
# iSNV Synonymous / Non-synonymous 분류
# ================================================================
# 입력 파일 3개:
#   [1] raw data 파일       (각 샘플 탭, ref_genome 포함)
#   [2] iSNV summary 파일   (단백질별 탭, Major/Minor 값)
#   [3] protein position 파일
#
# 출력 파일 2개:
#   [1] iSNV summary + Syn 주석 추가본
#       - 각 단백질 탭에 Ref_codon / Ref_AA / Codon# 컬럼 추가
#       - 각 샘플 Minor 값 옆에 Syn_Status 컬럼 추가
#   [2] iSNV count 파일 (NS / S 시트 채움)
# ================================================================

import subprocess
subprocess.run(['pip', 'install', 'biopython', '-q'], check=True)

from google.colab import files
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from Bio.Seq import Seq
import io, re

# ── 파일 업로드 ──────────────────────────────────────────────
print("▶ [1/3] raw data 파일을 업로드하세요")
print("   ※ 파일명에 'raw_data' 또는 'mapping_coverage' 포함된 파일")
print("      (각 샘플이 탭으로 구성, B열=number, C열=ref_genome)")
up1 = files.upload()
raw_file = list(up1.keys())[0]
print(f"  ✔ {raw_file}\n")

print("▶ [2/3] iSNV summary 파일을 업로드하세요")
print("   ※ 파일명에 'summary' 포함된 파일")
print("      (단백질별 탭: Spike/ORF1a/ORF1b 등, Major/Minor 값 포함)")
up2 = files.upload()
summary_file = list(up2.keys())[0]
print(f"  ✔ {summary_file}\n")

print("▶ [3/3] protein position 파일을 업로드하세요")
print("   ※ 파일명에 'protein_position' 포함된 파일")
up3 = files.upload()
protein_file = list(up3.keys())[0]
print(f"  ✔ {protein_file}\n")

# ================================================================
# 1. Reference genome 재구성 (raw data 첫 번째 탭에서 추출)
# ================================================================
print("▶ Reference genome 재구성 중...")
wb_raw = openpyxl.load_workbook(io.BytesIO(up1[raw_file]))
ws_ref = wb_raw[wb_raw.sheetnames[0]]

# number 컬럼(B)이 정수인 행만 수집, 헤더/수식 행은 건너뜀
ref_seq = {}
for r in range(2, ws_ref.max_row + 1):
    num  = ws_ref.cell(r, 2).value
    base = ws_ref.cell(r, 3).value
    if num is None or base is None:
        continue
    try:
        pos = int(num)
    except (ValueError, TypeError):
        continue
    b = str(base).strip().upper()
    if b in ('A', 'T', 'G', 'C'):   # ref base가 실제 염기인 행만
        ref_seq[pos] = b

print(f"  genome positions 수집: {len(ref_seq)}개")
if len(ref_seq) == 0:
    raise ValueError(
        "❌ ref_seq가 비어 있습니다.\n"
        "   [1/3]에 raw data 파일(B열=number, C열=ref_genome)을 업로드했는지 확인하세요.\n"
        "   iSNV summary 파일이나 다른 파일을 잘못 올린 경우 이 오류가 발생합니다."
    )

# ================================================================
# 2. Protein position 로드
# ================================================================
TARGET_PROTEINS = ['Spike', 'ORF1a', 'ORF1b', 'ORF3a', 'ORF6', 'ORF8', 'ORF10']

wb_prot = openpyxl.load_workbook(io.BytesIO(up3[protein_file]))
ws_prot = wb_prot.active

protein_ranges = {}
for row in ws_prot.iter_rows(values_only=True, min_row=2):
    name = str(row[0]).strip() if row[0] else ''
    if name in TARGET_PROTEINS:
        protein_ranges[name] = (int(row[1]), int(row[2]))

# CDS 시작 위치 (reading frame 기준점)
CDS_STARTS = {p: protein_ranges[p][0] for p in TARGET_PROTEINS if p in protein_ranges}

print("  단백질 CDS 시작점:")
for p, s in CDS_STARTS.items():
    print(f"    {p}: {s}")

# ================================================================
# 3. Synonymous 판단 함수
# ================================================================
def parse_allele(val):
    """'T 0.88' → 'T' / None → None"""
    if val is None:
        return None, None
    parts = str(val).strip().split()
    base = parts[0].upper() if parts else None
    freq = float(parts[1]) if len(parts) > 1 else None
    return base, freq

def classify_isnv(genome_pos, minor_base, protein):
    """
    반환: (syn_status, ref_codon, alt_codon, ref_aa, alt_aa, codon_num)
    syn_status: 'S' / 'NS' / 'Stop' / 'Intergenic'
    """
    cds_start = CDS_STARTS.get(protein)
    if cds_start is None:
        return 'Intergenic', '-', '-', '-', '-', '-'

    orf_pos    = genome_pos - cds_start       # 0-based CDS position
    if orf_pos < 0:
        return 'Intergenic', '-', '-', '-', '-', '-'

    codon_idx    = orf_pos // 3               # 0-based codon index
    codon_offset = orf_pos % 3               # position within codon (0,1,2)
    codon_start  = cds_start + codon_idx * 3  # genome pos of codon start

    ref_codon_bases = [ref_seq.get(codon_start + i, 'N') for i in range(3)]
    ref_codon = ''.join(ref_codon_bases)

    alt_codon_bases = ref_codon_bases.copy()
    alt_codon_bases[codon_offset] = minor_base.upper()
    alt_codon = ''.join(alt_codon_bases)

    ref_aa = str(Seq(ref_codon).translate())
    alt_aa = str(Seq(alt_codon).translate())

    if alt_aa == '*':
        status = 'Stop'
    elif ref_aa == alt_aa:
        status = 'S'
    else:
        status = 'NS'

    return status, ref_codon, alt_codon, ref_aa, alt_aa, codon_idx + 1

# ================================================================
# 4. iSNV summary 파일 로드 → 분류 → 새 파일 생성
# ================================================================
print("\n▶ iSNV summary 분류 시작...")

wb_in = openpyxl.load_workbook(io.BytesIO(up2[summary_file]))
wb_out = openpyxl.Workbook()
wb_out.remove(wb_out.active)

# 스타일 헬퍼
THIN   = Side(style='thin')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLOR_NS      = 'FADADD'   # 연분홍 - Non-synonymous
COLOR_S       = 'DAF0DA'   # 연초록 - Synonymous
COLOR_STOP    = 'FFD700'   # 금색   - Stop codon
COLOR_IG      = 'EEEEEE'   # 회색   - Intergenic
COLOR_ANNOT   = 'FFF2CC'   # 연노랑 - annotation 컬럼 헤더
COLOR_HDR     = 'BFBFBF'   # 헤더

PROTEIN_COLORS = {
    'Spike': 'D9E1F2', 'ORF1a': 'E2EFDA', 'ORF1b': 'FFF2CC',
    'ORF3a': 'FCE4D6', 'ORF6':  'EDEDED', 'ORF8':  'F4CCFF',
    'ORF10': 'CCF4FF', 'Other': 'F2F2F2',
}

STATUS_COLORS = {'NS': COLOR_NS, 'S': COLOR_S, 'Stop': COLOR_STOP, 'Intergenic': COLOR_IG}

def fill(color):
    return PatternFill("solid", fgColor=color)

def hcell(ws, r, c, val, color=COLOR_HDR, bold=True, center=True):
    cell = ws.cell(r, c, val)
    cell.font      = Font(bold=bold)
    cell.fill      = fill(color)
    cell.alignment = Alignment(horizontal='center' if center else 'left',
                                vertical='center', wrap_text=True)
    cell.border    = BORDER
    return cell

def dcell(ws, r, c, val, color=None, bold=False, center=True):
    cell = ws.cell(r, c, val)
    cell.font      = Font(bold=bold)
    if color:
        cell.fill  = fill(color)
    cell.alignment = Alignment(horizontal='center' if center else 'left',
                                vertical='center')
    cell.border    = BORDER
    return cell

# per-sample, per-protein iSNV count {sample: {protein: {S:0, NS:0}}}
sample_counts = {}

for sheet_name in wb_in.sheetnames:
    ws_src = wb_in[sheet_name]
    ws_dst = wb_out.create_sheet(title=sheet_name)

    pcolor = PROTEIN_COLORS.get(sheet_name, 'FFFFFF')
    is_target = sheet_name in TARGET_PROTEINS

    # ── 원본 헤더 파싱 ─────────────────────────────────────
    # Row1: Sample / sample_name / None / ...
    # Row2: Reference / Major / Minor / Major / Minor ...
    row1 = [ws_src.cell(1, c).value for c in range(1, ws_src.max_column + 1)]
    row2 = [ws_src.cell(2, c).value for c in range(1, ws_src.max_column + 1)]

    # 샘플 목록: (sample_name, major_src_col, minor_src_col) 1-based
    samples = []
    for c_idx, val in enumerate(row1, start=1):
        if val and val not in ('Sample',):
            samples.append({'name': val, 'maj': c_idx, 'min': c_idx + 1})

    # ── 출력 컬럼 설계 ─────────────────────────────────────
    # Col A  : genome position (Reference)
    # Col B  : Ref_codon
    # Col C  : Ref_AA
    # Col D  : Codon#
    # 이후: 샘플별로 Major(1) | Minor(1) | Syn_Status(1) = 3열씩
    # (Other 시트는 annotation 컬럼 추가하되 Syn=Intergenic)

    ANN_COLS = 3   # Ref_codon, Ref_AA, Codon#
    SAMPLE_WIDTH = 3  # Major, Minor, Syn_Status per sample

    # ── Row 1 헤더 ─────────────────────────────────────────
    # Col A
    ws_dst.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    hcell(ws_dst, 1, 1, 'Reference', color=COLOR_HDR)

    # Annotation cols B-D (merged 2 rows each)
    ann_headers = ['Ref_codon', 'Ref_AA', 'Codon#']
    for i, h in enumerate(ann_headers):
        col = 2 + i
        ws_dst.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
        hcell(ws_dst, 1, col, h, color=COLOR_ANNOT)

    # Sample headers
    for s_idx, sc in enumerate(samples):
        base_col = 2 + ANN_COLS + s_idx * SAMPLE_WIDTH
        # Row1: sample name merged across 3 cols
        ws_dst.merge_cells(start_row=1, start_column=base_col,
                           end_row=1, end_column=base_col + 2)
        hcell(ws_dst, 1, base_col, sc['name'], color=pcolor)
        # Row2: Major / Minor / Syn
        hcell(ws_dst, 2, base_col,     'Major',      color=pcolor)
        hcell(ws_dst, 2, base_col + 1, 'Minor',      color=pcolor)
        hcell(ws_dst, 2, base_col + 2, 'Syn_Status', color=COLOR_ANNOT)

    # ── 데이터 행 ──────────────────────────────────────────
    syn_count_sheet = {'S': 0, 'NS': 0}

    for r in range(3, ws_src.max_row + 1):
        src_row = [ws_src.cell(r, c).value for c in range(1, ws_src.max_column + 1)]
        genome_pos = src_row[0]
        if genome_pos is None:
            continue

        # annotation 계산 - 첫 번째 non-None minor allele 기준
        first_minor_base = None
        for sc in samples:
            if sc['min'] - 1 >= len(src_row):
                continue
            mval = src_row[sc['min'] - 1]
            if mval is not None:
                mb, _ = parse_allele(mval)
                if mb:
                    first_minor_base = mb
                    break

        if first_minor_base and is_target:
            status, ref_codon, alt_codon, ref_aa, alt_aa, codon_num = \
                classify_isnv(int(genome_pos), first_minor_base, sheet_name)
        else:
            status, ref_codon, alt_codon, ref_aa, alt_aa, codon_num = \
                ('Intergenic', '-', '-', '-', '-', '-')

        # Col A: genome position
        dcell(ws_dst, r, 1, genome_pos)

        # Annotation cols
        dcell(ws_dst, r, 2, ref_codon,  color=COLOR_ANNOT if ref_codon != '-' else None)
        dcell(ws_dst, r, 3, ref_aa,     color=COLOR_ANNOT if ref_aa    != '-' else None)
        dcell(ws_dst, r, 4, codon_num,  color=COLOR_ANNOT if codon_num != '-' else None)

        # 샘플별 데이터
        for s_idx, sc in enumerate(samples):
            base_col = 2 + ANN_COLS + s_idx * SAMPLE_WIDTH
            maj_val  = src_row[sc['maj'] - 1]
            min_val  = src_row[sc['min'] - 1]

            # Major
            dcell(ws_dst, r, base_col, maj_val)

            # Minor
            dcell(ws_dst, r, base_col + 1, min_val)

            # Syn_Status: per-sample minor allele로 재계산
            if min_val is not None:
                mb, _ = parse_allele(min_val)
                if mb and is_target:
                    st, _, _, _, _, _ = classify_isnv(int(genome_pos), mb, sheet_name)
                elif mb:
                    st = 'Intergenic'
                else:
                    st = None
            else:
                st = None

            if st:
                sc_color = STATUS_COLORS.get(st, None)
                dcell(ws_dst, r, base_col + 2, st, color=sc_color, bold=(st == 'NS'))

                # count 누적
                sname = sc['name']
                if sname not in sample_counts:
                    sample_counts[sname] = {p: {'S': 0, 'NS': 0} for p in TARGET_PROTEINS}
                if sheet_name in TARGET_PROTEINS and st in ('S', 'NS'):
                    sample_counts[sname][sheet_name][st] += 1

    # ── 컬럼 너비 ─────────────────────────────────────────
    ws_dst.column_dimensions['A'].width = 12
    ws_dst.column_dimensions['B'].width = 11  # Ref_codon
    ws_dst.column_dimensions['C'].width = 9   # Ref_AA
    ws_dst.column_dimensions['D'].width = 9   # Codon#
    for s_idx in range(len(samples)):
        base_col = 2 + ANN_COLS + s_idx * SAMPLE_WIDTH
        for offset, w in enumerate([12, 12, 13]):
            ws_dst.column_dimensions[get_column_letter(base_col + offset)].width = w

    # ── 행 높이 ───────────────────────────────────────────
    ws_dst.row_dimensions[1].height = 20
    ws_dst.row_dimensions[2].height = 18

    data_rows_count = ws_src.max_row - 2
    ns_cnt = sum(
        sample_counts.get(sc['name'], {}).get(sheet_name, {}).get('NS', 0)
        for sc in samples
    )
    s_cnt = sum(
        sample_counts.get(sc['name'], {}).get(sheet_name, {}).get('S', 0)
        for sc in samples
    )
    print(f"  [{sheet_name}] 완료 | NS={ns_cnt}, S={s_cnt}")

# ================================================================
# 5. Count 파일 생성 (NS / S 시트)
# ================================================================
print("\n▶ iSNV count 파일 생성 중...")

wb_count = openpyxl.load_workbook(io.BytesIO(up2[summary_file]))
# iSNV_NS, iSNV_S 시트 없으면 생성, 있으면 초기화
for sh_name in ['iSNV_NS', 'iSNV_S']:
    if sh_name in wb_count.sheetnames:
        del wb_count[sh_name]

# summary 파일의 샘플 목록 (Spike 시트 기준)
ws_spike = wb_count['Spike'] if 'Spike' in wb_count.sheetnames else wb_count[wb_count.sheetnames[0]]
row1_spike = [ws_spike.cell(1, c).value for c in range(1, ws_spike.max_column + 1)]
all_samples = [v for v in row1_spike if v and v != 'Sample']

def extract_group(sname):
    m = re.search(r'_(\d+\.?\d*)_', sname)
    return m.group(1) if m else 'unknown'

def group_sort(sname):
    try:    return float(extract_group(sname))
    except: return 999

sorted_samples = sorted(all_samples, key=group_sort)

HEADER_COLOR  = 'D6DCE4'
ROW_COLORS    = {'0.01': 'EAF4FF', '0.1': 'FFF4EA', '1': 'F4FFEA'}

for mode in ['NS', 'S']:
    ws_new = wb_count.create_sheet(title=f'iSNV_{mode}')

    headers = ['Sample', 'Group', f'total_iSNVs_{mode}'] + \
              [f'{p}_iSNVs' for p in TARGET_PROTEINS]

    for c_idx, h in enumerate(headers, start=1):
        hcell(ws_new, 1, c_idx, h, color=HEADER_COLOR)

    total_row_sums = [0] * (len(TARGET_PROTEINS) + 1)

    for r_off, sname in enumerate(sorted_samples):
        dr = 2 + r_off
        group = extract_group(sname)
        row_color = ROW_COLORS.get(group, 'FFFFFF')

        per_prot = sample_counts.get(sname, {p: {'S': 0, 'NS': 0} for p in TARGET_PROTEINS})
        vals = [per_prot.get(p, {}).get(mode, 0) for p in TARGET_PROTEINS]
        total = sum(vals)

        row_vals = [sname, group, total] + vals
        for c_idx, val in enumerate(row_vals, start=1):
            dcell(ws_new, dr, c_idx, val, color=row_color)

        total_row_sums[0] += total
        for i, v in enumerate(vals):
            total_row_sums[i + 1] += v

    # Sum 행
    sum_row = 2 + len(sorted_samples)
    dcell(ws_new, sum_row, 1, 'Sum',  color=COLOR_HDR, bold=True)
    dcell(ws_new, sum_row, 2, '',     color=COLOR_HDR)
    for c_idx, val in enumerate([total_row_sums[0]] + total_row_sums[1:], start=3):
        dcell(ws_new, sum_row, c_idx, val, color=COLOR_HDR, bold=True)

    # 컬럼 너비
    ws_new.column_dimensions['A'].width = 25
    ws_new.column_dimensions['B'].width = 10
    for c_idx in range(3, len(headers) + 1):
        ws_new.column_dimensions[get_column_letter(c_idx)].width = 14

    print(f"  [iSNV_{mode}] 완료 | 총 {total_row_sums[0]}개")

# ================================================================
# 6. 저장 및 다운로드
# ================================================================
base = re.sub(r'\.xlsx$', '', summary_file, flags=re.IGNORECASE)

out1 = base + '_Syn분류.xlsx'
out2 = base + '_iSNV_count_NS_S.xlsx'

for out_name, wb_obj in [(out1, wb_out), (out2, wb_count)]:
    buf = io.BytesIO()
    wb_obj.save(buf)
    buf.seek(0)
    with open(out_name, 'wb') as f:
        f.write(buf.read())
    files.download(out_name)
    print(f"\n✅ 저장: {out_name}")

print("\n🎉 모든 작업 완료!")